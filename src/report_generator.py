#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import json
import shutil
import logging
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional
import openpyxl.cell.cell
import traceback
import xlrd
import re
from copy import copy
import openpyxl.styles

# 确保日志目录存在
logs_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logs')
os.makedirs(logs_dir, exist_ok=True)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(logs_dir, 'template_report.log'), encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# 导入精确提取函数
from src.extract_excel_cell import extract_product_info_direct, extract_cell_value


def load_json_file(file_path: str) -> Dict[str, Any]:
    """
    加载JSON文件

    Args:
        file_path: JSON文件路径

    Returns:
        Dict[str, Any]: 加载的JSON数据
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logger.info(f"成功加载JSON文件: {file_path}")
        return data
    except Exception as e:
        logger.error(f"加载JSON文件时出错: {e}")
        raise


def extract_product_info(excel_file: str) -> Dict[str, str]:
    """
    从Excel文件中精确提取产品信息（产品编码、图号、材料）

    Args:
        excel_file: Excel文件路径

    Returns:
        Dict[str, str]: 产品信息字典
    """
    product_info = {
        "product_code": "",  # 产品编码
        "drawing_no": "",  # 图号
        "material": "",  # 材料
        "version": "",  # 版本号
        "part_name": ""  # 零件名称
    }
    logger.info(f"开始提取产品信息: {excel_file}")

    try:
        # 从指定单元格精确提取数据
        product_info["product_code"] = extract_cell_value(excel_file, 'D4')
        product_info["drawing_no"] = extract_cell_value(excel_file, 'C5')
        product_info["material"] = extract_cell_value(excel_file, 'H6')
        product_info["version"] = extract_cell_value(excel_file, 'A4')
        product_info["part_name"] = extract_cell_value(excel_file, 'C6')

        logger.info(f"从D4单元格提取到产品编码: {product_info['product_code']}")
        logger.info(f"从C5单元格提取到图号: {product_info['drawing_no']}")
        logger.info(f"从H6单元格提取到材料: {product_info['material']}")
        logger.info(f"从A4单元格提取到版本号: {product_info['version']}")
        logger.info(f"从C6单元格提取到零件名称: {product_info['part_name']}")

        # 处理产品编码，从"产品编码:21013554"中提取"21013554"
        if "：" in product_info["product_code"]:
            parts = product_info["product_code"].split("：", 1)
            if len(parts) > 1:
                product_info["product_code"] = parts[1].strip()
        elif ":" in product_info["product_code"]:
            parts = product_info["product_code"].split(":", 1)
            if len(parts) > 1:
                product_info["product_code"] = parts[1].strip()

        # 处理版本号，从"版本:03"中提取"03"
        if "：" in product_info["version"]:
            parts = product_info["version"].split("：", 1)
            if len(parts) > 1:
                product_info["version"] = parts[1].strip()
        elif ":" in product_info["version"]:
            parts = product_info["version"].split(":", 1)
            if len(parts) > 1:
                product_info["version"] = parts[1].strip()

    except Exception as e:
        logger.error(f"从Excel精确提取产品信息时出错: {e}")

        # 如果精确提取失败，可以添加备用方案
        # 比如使用关键词搜索等

    return product_info


def extract_process_info_direct(excel_file: str) -> List[Dict[str, str]]:
    """
    直接从Excel文件中提取工序信息，只处理第一个工作表

    Args:
        excel_file: Excel文件路径

    Returns:
        List[Dict[str, str]]: 工序信息列表，格式为[{process_name, process_code, process_desc}, ...]
    """
    # 检查文件是否为JSON文件（新结构）
    if excel_file.lower().endswith('.json'):
        return extract_process_info_from_json(excel_file)

    process_info = []

    # 判断文件格式
    file_extension = os.path.splitext(excel_file)[1].lower()

    try:
        # 根据文件扩展名选择合适的库
        if file_extension == '.xlsx':
            # 使用openpyxl处理.xlsx文件
            wb = load_workbook(excel_file, data_only=True)
            if len(wb.sheetnames) == 0:
                logger.error(f"Excel文件 {excel_file} 中没有工作表")
                return []

            # 只处理第一个工作表
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]

            logger.info(f"仅处理第一个工作表: {sheet_name}")

            # 查找工序标题行
            header_row = None
            for row_idx in range(1, min(50, ws.max_row + 1)):  # 只搜索前50行
                if ws.cell(row=row_idx, column=1).value == "工序":
                    header_row = row_idx
                    break

            if header_row is None:
                logger.warning(f"工作表 {sheet_name} 中未找到工序标题行")
                return []

            # 从工序标题行之后提取工序信息
            for row_idx in range(header_row + 1, ws.max_row + 1):
                process_name = ws.cell(row=row_idx, column=1).value

                # 跳过空行或不含工序名称的行
                if not process_name or str(process_name).strip() == "":
                    continue

                # 获取工序代码和描述
                process_code = None
                process_desc = None

                # 根据不同工作表格式提取信息
                if sheet_name == "04":
                    process_code = ws.cell(row=row_idx, column=2).value
                    process_desc = ws.cell(row=row_idx, column=3).value
                elif sheet_name in ["03", "02", "工艺"]:
                    process_code = ws.cell(row=row_idx, column=2).value
                    process_desc = ws.cell(row=row_idx, column=2).value

                # 记录工序信息
                process_info.append({
                    "process_name": str(process_name).strip() if process_name else "",
                    "process_code": str(process_code).strip() if process_code else "",
                    "process_desc": str(process_desc).strip() if process_desc else ""
                })

                logger.info(f"提取到工序: {process_name}, 工序代码: {process_code}, 工艺说明: {process_desc}")

        elif file_extension == '.xls':
            # 使用xlrd处理.xls文件
            wb = xlrd.open_workbook(excel_file)
            if wb.nsheets == 0:
                logger.error(f"Excel文件 {excel_file} 中没有工作表")
                return []

            # 只处理第一个工作表
            ws = wb.sheet_by_index(0)
            sheet_name = ws.name

            logger.info(f"仅处理第一个工作表: {sheet_name}")

            # 查找工序标题行
            header_row = None
            for row_idx in range(min(50, ws.nrows)):  # 只搜索前50行
                if ws.cell_value(row_idx, 0) == "工序":
                    header_row = row_idx
                    break

            if header_row is None:
                logger.warning(f"工作表 {sheet_name} 中未找到工序标题行")
                return []

            # 从工序标题行之后提取工序信息
            for row_idx in range(header_row + 1, ws.nrows):
                process_name = ws.cell_value(row_idx, 0)

                # 跳过空行或不含工序名称的行
                if not process_name or str(process_name).strip() == "":
                    continue

                # 获取工序代码和描述
                process_code = None
                process_desc = None

                # 根据不同工作表格式提取信息
                if sheet_name == "04":
                    process_code = ws.cell_value(row_idx, 1) if ws.ncols > 1 else ""
                    process_desc = ws.cell_value(row_idx, 2) if ws.ncols > 2 else ""
                elif sheet_name in ["03", "02", "工艺"]:
                    process_code = ws.cell_value(row_idx, 1) if ws.ncols > 1 else ""
                    process_desc = ws.cell_value(row_idx, 1) if ws.ncols > 1 else ""

                # 记录工序信息
                process_info.append({
                    "process_name": str(process_name).strip() if process_name else "",
                    "process_code": str(process_code).strip() if process_code else "",
                    "process_desc": str(process_desc).strip() if process_desc else ""
                })

                logger.info(f"提取到工序: {process_name}, 工序代码: {process_code}, 工艺说明: {process_desc}")

        else:
            logger.error(f"不支持的文件格式: {file_extension}")
            return []

    except Exception as e:
        logger.error(f"从Excel文件 {excel_file} 中提取工序信息时出错: {e}")
        logger.error(traceback.format_exc())

    logger.info(f"共提取到 {len(process_info)} 个工序信息")
    return process_info


def extract_process_info_from_json(json_file: str) -> List[Dict[str, str]]:
    """
    从JSON文件中提取工序信息

    Args:
        json_file: JSON文件路径

    Returns:
        List[Dict[str, str]]: 工序信息列表，格式为[{process_name, process_code, process_desc}, ...]
    """
    process_info = []

    try:
        # 加载JSON文件
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 提取工序信息
        if 'processes' in data:
            process_info = data['processes']
            logger.info(f"从JSON文件 {json_file} 中提取到 {len(process_info)} 个工序")
        else:
            logger.warning(f"JSON文件 {json_file} 中未找到工序信息")

    except Exception as e:
        logger.error(f"从JSON文件 {json_file} 中提取工序信息时出错: {e}")

    return process_info


def get_appearance_requirements(process_list: List[Dict[str, str]], appearance_map: Dict[str, List[Dict[str, Any]]]) -> \
List[str]:
    """
    根据工序列表和外观要求对照表生成检验项列表

    Args:
        process_list: 工序信息列表
        appearance_map: 外观要求对照表

    Returns:
        List[str]: 外观要求检验项目列表
    """
    inspection_items = []

    # 使用集合记录已处理的工序名称，避免重复处理
    processed_process_names = set()

    # 遍历工序列表
    for process in process_list:
        # 获取工序名称和代码
        process_name = process.get("process_name", "")
        process_code = process.get("process_code", "")

        if not process_name:
            continue

        # 如果该工序名称已经处理过，则跳过
        if process_name in processed_process_names:
            continue

        # 在外观要求对照表中查找匹配项
        if process_name in appearance_map:
            requirements = appearance_map[process_name]
            logger.info(f"工序 {process_name} 在外观要求对照表中找到 {len(requirements)} 个匹配项")

            # 添加该工序的所有外观要求
            for req in requirements:
                req_desc = req.get("description", "")

                # 格式化外观要求，添加工序名称和编号
                formatted_req = f"外观：{process_name}({process_code})-{req_desc}"
                inspection_items.append(formatted_req)
                logger.info(f"添加外观要求: {formatted_req}")

            # 将该工序名称添加到已处理集合
            processed_process_names.add(process_name)
        else:
            logger.info(f"工序 {process_name} 在外观要求对照表中未找到匹配项，跳过")

    return inspection_items


def clean_cad_text(text: str) -> str:
    """
    优雅地处理CAD文本格式，转换为人类可读的格式

    Args:
        text: 原始CAD文本，如 \A1;96-%%c4.6{\H0.7x;\S+0.1^ 0;}

    Returns:
        str: 处理后的文本，如 φ4.6±(0.1, 0)
    """
    if not text:
        return ""

    # 1. 提取关键数据：寻找圆直径和公差信息
    diameter_match = re.search(r'%%c(\d+(\.\d+)?)', text)
    tolerance_match = re.search(r'\\S\+([0-9.]+)\^\s*([0-9.]+)', text)

    # 如果找到直径
    if diameter_match:
        diameter = diameter_match.group(1)

        # 处理公差
        if tolerance_match:
            positive_tol = tolerance_match.group(1)
            negative_tol = tolerance_match.group(2)
            return f"φ{diameter}±({positive_tol}, {negative_tol})"
        else:
            # 如果没有公差信息，使用默认公差
            return f"φ{diameter}±0.2"

    # 如果没有直径但有数字，可能是普通尺寸
    else:
        # 清理文本，去除格式标记
        cleaned = re.sub(r'\\[A-Za-z][^;]*;', '', text)  # 移除 \A1; 等标记
        cleaned = re.sub(r'{\\[^}]*}', '', cleaned)  # 移除花括号及其内容
        cleaned = re.sub(r'{[^}]*}', '', cleaned)

        # 替换标准符号
        cleaned = cleaned.replace('%%P', '±')
        cleaned = cleaned.replace('%%c', 'φ')

        # 提取数字部分
        number_match = re.search(r'(\d+(\.\d+)?)', cleaned)
        if number_match:
            # 如果已包含公差信息
            if '±' in cleaned:
                return cleaned
            # 否则添加默认公差
            else:
                number = number_match.group(1)
                return re.sub(r'(\d+(\.\d+)?)', f'\\1±0.2', cleaned, count=1)

        return cleaned


def extract_dwg_text(dwg_data: Dict[str, Any]) -> List[str]:
    """
    从DWG数据中提取MTEXT实体的文本内容，处理格式并添加编号

    Args:
        dwg_data: DWG数据

    Returns:
        List[str]: 包含处理后的尺寸文本列表
    """
    inspection_items = []

    # 提取MTEXT实体
    for idx, entity in enumerate(dwg_data.get('mtext', []), start=1):
        text = entity.get('text', '')
        if text and len(text) > 1:  # 过滤掉过短的文本
            # 使用优雅的清理函数处理文本
            cleaned_text = clean_cad_text(text)

            # 拼接带编号的"尺寸"前缀
            dimension_text = f"尺寸{idx}：{cleaned_text.strip()}"
            inspection_items.append(dimension_text)
            logger.info(f"提取到DWG文本: {dimension_text}")

    return inspection_items


def parse_material_info(material_str: str) -> Dict[str, str]:
    """
    解析材料信息，将其分割为材质和料厚

    Args:
        material_str: 材料信息字符串，如"Fe,08 T=3"或"Fe,08 T3"

    Returns:
        Dict[str, str]: 包含材质和料厚的字典
    """
    result = {
        "material_type": "",  # 材质
        "thickness": "",  # 料厚
        "thickness_value": 0.0  # 料厚数值
    }

    if not material_str:
        return result

    # 按空格分割
    parts = material_str.split(" ")

    # 提取材质
    if len(parts) > 0:
        result["material_type"] = parts[0].strip()

    # 提取料厚
    if len(parts) > 1:
        thickness_part = parts[1].strip()
        result["thickness"] = thickness_part

        # 提取料厚的数值部分，支持T=3和T3两种格式
        thickness_match = re.search(r'T=?(\d+(\.\d+)?)', thickness_part, re.IGNORECASE)
        if thickness_match:
            try:
                result["thickness_value"] = float(thickness_match.group(1))
            except ValueError:
                logger.warning(f"无法将料厚转换为浮点数: {thickness_match.group(1)}")
        else:
            # 如果没有匹配到标准格式，尝试直接提取数字
            numeric_match = re.search(r'(\d+(\.\d+)?)', thickness_part)
            if numeric_match:
                try:
                    result["thickness_value"] = float(numeric_match.group(1))
                    # 如果厚度前没有T标记，则添加
                    if not thickness_part.upper().startswith('T'):
                        result["thickness"] = f"T{result['thickness_value']}"
                except ValueError:
                    logger.warning(f"无法将料厚转换为浮点数: {numeric_match.group(1)}")

    logger.info(
        f"解析材料信息: {material_str} -> 材质: {result['material_type']}, 料厚: {result['thickness']}, 料厚数值: {result['thickness_value']}")
    return result


def get_thickness_tolerance(thickness_value: float, thickness_map_file: str) -> Dict[str, float]:
    """
    根据厚度查找对应的允许偏差

    Args:
        thickness_value: 厚度值
        thickness_map_file: 厚度公差映射文件路径

    Returns:
        Dict[str, float]: 包含正公差和负公差的字典
    """
    result = {
        "tolerance_str": "±0.00",
        "upper_tolerance": 0.0,
        "lower_tolerance": 0.0
    }

    if thickness_value <= 0:
        return result

    try:
        # 加载厚度公差映射
        with open(thickness_map_file, 'r', encoding='utf-8') as f:
            thickness_tolerance = json.load(f)

        # 查找对应的公差
        for range_str, tolerance in thickness_tolerance.items():
            # 提取范围的下限和上限
            range_match = re.search(r'>(\d+\.\d+)~(\d+\.\d+)', range_str)
            if range_match:
                lower = float(range_match.group(1))
                upper = float(range_match.group(2))

                # 检查厚度是否在范围内
                if lower < thickness_value <= upper:
                    result["tolerance_str"] = tolerance

                    # 提取公差值
                    tolerance_match = re.search(r'±(\d+\.\d+)', tolerance)
                    if tolerance_match:
                        tolerance_value = float(tolerance_match.group(1))
                        result["upper_tolerance"] = tolerance_value
                        result["lower_tolerance"] = tolerance_value

                    logger.info(f"厚度 {thickness_value} 对应的公差: {tolerance}")
                    break
    except Exception as e:
        logger.error(f"获取厚度公差时出错: {e}")
        logger.error(traceback.format_exc())

    return result


def insert_drawing_image(report_path: str, image_path: str) -> None:
    """
    将图纸图像插入到报告的第二个sheet的A1单元格

    Args:
        report_path: Excel报告文件路径
        image_path: 图像文件路径
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage

        # 加载工作簿
        wb = load_workbook(report_path)

        # 获取或创建第二个sheet
        if len(wb.sheetnames) < 2:
            sheet = wb.create_sheet(title="图纸")
        else:
            sheet = wb.worksheets[1]

        # 设置sheet名称为"图纸"
        sheet.title = "图纸"

        # 调整单元格大小以适应A3图像
        # sheet.column_dimensions['A'].width = 120  # 约等于297mm
        # sheet.row_dimensions[1].height = 420  # 约等于420mm

        # 设置单元格行高和列宽与打印尺寸一致
        def extract_resolution(s):
            pattern = r'(\d+\.\d{2})_x_(\d+\.\d{2})'
            match = re.search(pattern, s)
            if match:
                return match.groups()
            else:
                raise ValueError("字符串格式不符合要求")

        from src.export_image import CADConfig
        config = CADConfig()
        width, height = extract_resolution(config.media_name)
        sheet.column_dimensions['A'].width = width
        sheet.row_dimensions[1].height = height

        # 设置单元格背景为黑色
        from openpyxl.styles import PatternFill
        black_fill = PatternFill(
            start_color='000000',  # 黑色十六进制代码
            end_color='000000',
            fill_type='solid'
        )

        # 设置A1单元格背景为黑色
        sheet['A1'].fill = black_fill

        # 创建图像对象
        img = XLImage(image_path)

        # 将图像插入到A1单元格
        sheet.add_image(img, 'A1')

        # 保存工作簿
        wb.save(report_path)
        logger.info(f"成功将图像插入到报告: {report_path}")

    except Exception as e:
        logger.error(f"插入图像到Excel报告时出错: {e}")
        logger.error(traceback.format_exc())
        raise


def generate_template_report(dwg_file: str, excel_file: str, appearance_map_file: str, template_file: str,
                             output_dir: str = 'outputs', original_excel_file: str = None) -> str:
    """
    根据模板生成检验报告

    Args:
        dwg_file: DWG数据文件路径
        excel_file: Excel数据文件路径或原始Excel文件路径
        appearance_map_file: 外观要求对照表文件路径
        template_file: 报告模板文件路径
        output_dir: 输出目录
        original_excel_file: 原始Excel文件路径（可选）

    Returns:
        str: 生成的报告文件路径
    """
    try:
        # 创建输出目录
        os.makedirs(output_dir, exist_ok=True)

        # 检查输入文件路径
        if not os.path.exists(dwg_file):
            raise FileNotFoundError(f"DWG数据文件不存在: {dwg_file}")
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"Excel文件不存在: {excel_file}")
        if not os.path.exists(appearance_map_file):
            raise FileNotFoundError(f"外观要求对照表文件不存在: {appearance_map_file}")
        if not os.path.exists(template_file):
            raise FileNotFoundError(f"报告模板文件不存在: {template_file}")

        # 加载数据
        logger.info(f"加载DWG数据: {dwg_file}")
        dwg_data = load_json_file(dwg_file)

        # 检查excel_file是JSON文件还是原始Excel文件
        is_json_file = excel_file.lower().endswith('.json')

        if is_json_file:
            # 如果是JSON文件，加载Excel数据
            logger.info(f"加载Excel JSON数据: {excel_file}")
            excel_data = load_json_file(excel_file)

            # 优先使用传入的原始Excel文件路径
            if original_excel_file and os.path.exists(original_excel_file):
                logger.info(f"使用传入的原始Excel文件: {original_excel_file}")
                original_excel_file = original_excel_file
            else:
                # 尝试查找原始Excel文件
                original_excel_filename = os.path.basename(excel_file).replace('_excel.json', '.xlsx')
                if not original_excel_filename.endswith(('.xlsx', '.xls')):
                    original_excel_filename += '.xlsx'

                # 尝试在几个可能的位置查找原始Excel文件
                possible_locations = [
                    os.path.join(os.path.dirname(excel_file), original_excel_filename),
                    os.path.join('uploads', 'excel', original_excel_filename),
                    os.path.join('uploads', original_excel_filename)
                ]

                original_excel_file = None
                for location in possible_locations:
                    if os.path.exists(location):
                        original_excel_file = location
                        logger.info(f"找到原始Excel文件: {original_excel_file}")
                        break
        else:
            # 如果是原始Excel文件，直接使用
            logger.info(f"使用原始Excel文件: {excel_file}")
            original_excel_file = excel_file

            # 为了兼容性，生成一个临时的Excel数据结构
            from src.excel_parser import parse_excel_file
            excel_data = parse_excel_file(excel_file)

        logger.info(f"加载外观要求对照表: {appearance_map_file}")
        appearance_map = load_json_file(appearance_map_file)

        # 提取基本信息 - 使用精确提取函数
        if original_excel_file:
            logger.info(f"使用精确提取函数从原始Excel文件中提取产品信息")
            product_info = extract_product_info(original_excel_file)
        else:
            logger.warning(f"未找到原始Excel文件，将使用JSON数据提取产品信息")
            product_info = extract_product_info(excel_file)

        logger.info(f"提取的产品信息: {product_info}")

        # 获取DWG文件名并生成输出文件名
        dwg_file_name = dwg_data.get('file_name', '')
        base_name = os.path.splitext(dwg_file_name)[0]
        output_file = os.path.join(output_dir, f"{base_name}_report.xlsx")

        # 首先复制模板文件创建报告
        logger.info(f"复制模板文件 {template_file} 到 {output_file}")
        shutil.copy2(template_file, output_file)

        # 处理DWG转换和图像插入
        dxf_path = os.path.join('outputs', f"{base_name}.dxf")
        if not os.path.exists(dxf_path):
            # 从原始Excel文件路径中提取任务ID
            task_id = None
            if original_excel_file:
                # 从路径中提取任务ID
                path_parts = original_excel_file.split(os.sep)
                for part in path_parts:
                    if len(part) == 36 and '-' in part:  # UUID格式
                        task_id = part
                        break

            # 构建正确的DWG文件路径
            if task_id:
                dwg_path = os.path.join('uploads', task_id, dwg_file_name)
            else:
                dwg_path = os.path.join('uploads', dwg_file_name)

            from src.dwg_parser import convert_dwg_to_dxf
            dxf_path = convert_dwg_to_dxf(dwg_path, dxf_path)

        if dxf_path and os.path.exists(dxf_path):
            # 转换DXF为图像
            # from src.dxf_to_image import convert_dxf_to_image
            # image_path = os.path.join('outputs', f"{base_name}.png")
            # image_path = convert_dxf_to_image(dxf_path, image_path)
            filename = os.path.basename(dxf_path)
            _dwg_file = filename.rsplit(".", 1)[0] + ".dwg"
            project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            test_file_dir = [os.path.join(project_dir, 'test')]
            input_file = os.path.join(test_file_dir[0], _dwg_file)

            from src.export_image import process_dwg
            retval = process_dwg(input_file, os.path.join(project_dir, "outputs"))
            image_path = os.path.join('outputs', f"{base_name}.png")
            if retval and os.path.exists(image_path):
                # 将图像插入到报告中
                insert_drawing_image(output_file, image_path)

        # 使用openpyxl加载工作簿
        wb = load_workbook(output_file)
        ws = wb.active

        # 安全设置单元格值的函数（处理合并单元格）
        def set_cell_value(sheet, coord, value):
            try:
                cell = sheet[coord]
                # 检查是否为合并单元格
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # 找出此合并单元格区域
                    for merged_range in sheet.merged_cells.ranges:
                        if coord in merged_range:
                            # 获取合并区域的左上角单元格
                            min_col, min_row, max_col, max_row = merged_range.bounds
                            # 设置左上角单元格的值
                            sheet.cell(row=min_row, column=min_col).value = value
                            break
                else:
                    cell.value = value
            except Exception as e:
                logger.error(f"设置单元格 {coord} 值时出错: {e}")
                # 尝试直接通过行列设置
                col_letter = coord[0]
                row_num = int(coord[1:])
                col_idx = ord(col_letter.upper()) - ord('A') + 1
                sheet.cell(row=row_num, column=col_idx).value = value

        # 1. 填充物料编码到D2合并单元格
        logger.info("填充物料编码")
        set_cell_value(ws, 'D2', product_info['product_code'])

        # 2. 填充图号/名称到J2单元格，将图号与零件名称拼接
        logger.info("填充图号/零件名称")
        drawing_part_name = product_info['drawing_no']
        if product_info['part_name']:
            drawing_part_name = f"{product_info['drawing_no']}/{product_info['part_name']}"
        set_cell_value(ws, 'J2', drawing_part_name)

        # 3. 填充版本号到Q2单元格
        logger.info("填充版本号")
        set_cell_value(ws, 'Q2', product_info['version'])

        # 4. 填充材料/料厚到B4单元格
        logger.info("填充材料")
        set_cell_value(ws, 'D4', product_info['material'])

        # 5. 解析材料信息，获取材质和料厚
        logger.info("解析材料信息")
        material_info = parse_material_info(product_info['material'])

        # 创建材质和料厚的检验项目
        material_type_item = f"材质：{material_info['material_type']}"
        thickness_item = f"料厚：{material_info['thickness']}"

        # 查找厚度对应的公差
        thickness_map_file = os.path.join('maps', 'thickness_map.json')
        if os.path.exists(thickness_map_file) and material_info['thickness_value'] > 0:
            logger.info(f"查找厚度 {material_info['thickness_value']} 对应的公差")
            thickness_tolerance = get_thickness_tolerance(material_info['thickness_value'], thickness_map_file)

            # 更新料厚项目，加入公差信息
            thickness_item = f"料厚：{material_info['thickness']} {thickness_tolerance['tolerance_str']}"

        # 6. 获取工序信息和对应的外观要求
        logger.info("提取工序信息")
        process_list = extract_process_info_direct(excel_file)
        logger.info(f"找到 {len(process_list)} 个工序")
        logger.info(f"工序信息: {process_list}")

        logger.info("匹配外观要求")
        inspection_items = get_appearance_requirements(process_list, appearance_map)
        logger.info(f"找到 {len(inspection_items)} 个外观要求项")

        # 添加工艺流程卡中的工序描述到检验项目
        logger.info("添加工序描述到检验项目")
        # 使用集合记录已添加的工序描述，避免重复
        added_process_descs = set()

        for process in process_list:
            process_name = process.get("process_name", "")
            process_code = process.get("process_code", "")
            process_desc = process.get("process_desc", "")

            # 如果工序描述不为空，且未添加过，则添加为检验项
            if process_desc and process_desc.strip() != "":
                # 创建唯一标识，用于检查重复
                desc_identifier = f"{process_name}-{process_code}-{process_desc}"

                if desc_identifier not in added_process_descs:
                    desc_item = f"加工：{process_name}({process_code})-{process_desc}"
                    inspection_items.append(desc_item)
                    logger.info(f"添加工序描述: {desc_item}")

                    # 添加到已处理集合
                    added_process_descs.add(desc_identifier)
                else:
                    logger.info(f"跳过重复的工序描述: {process_name}({process_code})-{process_desc}")

        # 7. 从DWG中提取MTEXT实体
        logger.info("提取DWG文本")
        logger.info(f"DWG数据: {dwg_data}")
        dwg_text_items = extract_dwg_text(dwg_data)
        logger.info(f"找到 {len(dwg_text_items)} 个DWG文本项")
        logger.info(f"DWG文本: {dwg_text_items}")

        # 8. 将DWG中的文本添加到检验项目列表
        inspection_items.extend(dwg_text_items)

        logger.info(f"外观要求: {inspection_items}")
        # 9. 填充检验项目到Excel
        logger.info("填充检验项目到Excel")

        # 设置起始行为7，结束行为40（共34行可用空间）
        start_row = 7
        end_row = 40
        max_items = end_row - start_row + 1  # 最多可填充34个项目
        logger.info(f"设置检验项目起始行: {start_row}，结束行: {end_row}，最多可填充 {max_items} 个项目")

        # 先填充材质和料厚的检验项目（第7行和第8行）
        # 第7行：材质
        set_cell_value(ws, f"A{start_row}", 1)
        set_cell_value(ws, f"B{start_row}", material_type_item)
        set_cell_value(ws, f"H{start_row}", "性能")

        # 第8行：料厚
        set_cell_value(ws, f"A{start_row + 1}", 2)
        set_cell_value(ws, f"B{start_row + 1}", thickness_item)
        set_cell_value(ws, f"H{start_row + 1}", "尺寸")

        # 供应商判定结果
        set_cell_value(ws, f"L{start_row}", "□OK  □NG")
        ws[f"L{start_row}"].font = openpyxl.styles.Font(name="宋体")
        set_cell_value(ws, f"P{start_row}", "□OK  □NG")
        ws[f"P{start_row}"].font = openpyxl.styles.Font(name="宋体")

        # 填充料厚的公差范围到F和G列
        if material_info['thickness_value'] > 0 and 'thickness_tolerance' in locals() and thickness_tolerance[
            'upper_tolerance'] > 0:
            # 计算上下限
            base_value = material_info['thickness_value']
            upper_tolerance = thickness_tolerance['upper_tolerance']
            lower_tolerance = thickness_tolerance['lower_tolerance']

            # F列：规格下限值
            lower_limit = base_value - lower_tolerance
            set_cell_value(ws, f"F{start_row + 1}", f"{lower_limit:.2f}")

            # G列：规格上限值
            upper_limit = base_value + upper_tolerance
            set_cell_value(ws, f"G{start_row + 1}", f"{upper_limit:.2f}")

        set_cell_value(ws, f"L{start_row + 1}", "□OK  □NG")
        ws[f"L{start_row + 1}"].font = openpyxl.styles.Font(name="宋体")
        set_cell_value(ws, f"P{start_row + 1}", "□OK  □NG")
        ws[f"P{start_row + 1}"].font = openpyxl.styles.Font(name="宋体")

        # 如果检验项目数量超过可用行数（减去已使用的2行），进行截断
        available_rows = max_items - 2
        if len(inspection_items) > available_rows:
            logger.warning(f"检验项目数量 {len(inspection_items)} 超过最大可用行数 {available_rows}，将截断多余的项目")
            inspection_items = inspection_items[:available_rows]

        # 从第9行开始填充其他检验项目
        remaining_row_start = start_row + 2

        # 填充其他检验项目
        for i, item in enumerate(inspection_items, start=1):
            row_index = remaining_row_start + i - 1
            logger.debug(f"填充第 {i} 个检验项目到第 {row_index} 行: {item}")

            # A列：序号，从3开始（因为1和2已用于材质和料厚）
            set_cell_value(ws, f"A{row_index}", i + 2)

            # B列：检验项目
            set_cell_value(ws, f"B{row_index}", item)

            # H列：检验类别
            category = "性能"  # 默认值

            # 提取公差信息，计算上下限
            tolerance_match = re.search(r'(\d+(\.\d+)?)±(\d+(\.\d+)?)', item)
            if tolerance_match:
                base_value = float(tolerance_match.group(1))
                tolerance = float(tolerance_match.group(3))

                # F列：规格下限值
                lower_limit = base_value - tolerance
                set_cell_value(ws, f"F{row_index}", f"{lower_limit:.1f}")

                # G列：规格上限值
                upper_limit = base_value + tolerance
                set_cell_value(ws, f"G{row_index}", f"{upper_limit:.1f}")

            # 更复杂的公差格式 ±(a, b)
            complex_tolerance_match = re.search(r'(\d+(\.\d+)?)±\((\d+(\.\d+)?),\s*(\d+(\.\d+)?)\)', item)
            if complex_tolerance_match and not tolerance_match:
                base_value = float(complex_tolerance_match.group(1))
                upper_tolerance = float(complex_tolerance_match.group(3))
                lower_tolerance = float(complex_tolerance_match.group(5))

                # F列：规格下限值
                lower_limit = base_value - lower_tolerance
                set_cell_value(ws, f"F{row_index}", f"{lower_limit:.1f}")

                # G列：规格上限值
                upper_limit = base_value + upper_tolerance
                set_cell_value(ws, f"G{row_index}", f"{upper_limit:.1f}")

            if "外观" in item:
                category = "外观"
            elif "尺寸" in item:
                category = "尺寸"
            elif "加工" in item:
                category = "性能"

            set_cell_value(ws, f"H{row_index}", category)

            # L列：供应商判定结果
            set_cell_value(ws, f"L{row_index}", "□OK  □NG")
            # 设置单元格字体为普通字体
            ws[f"L{row_index}"].font = openpyxl.styles.Font(name="宋体")

            # P列：判定结果
            set_cell_value(ws, f"P{row_index}", "□OK  □NG")
            # 设置单元格字体为普通字体
            ws[f"P{row_index}"].font = openpyxl.styles.Font(name="宋体")

        # 保存Excel文件
        logger.info(f"保存Excel文件: {output_file}")
        wb.save(output_file)
        logger.info(f"检验报告已生成: {output_file}")

        return output_file
    except Exception as e:
        logger.error(f"生成模板报告时出错: {e}")
        logger.error(traceback.format_exc())
        raise


# 用于测试的代码
if __name__ == "__main__":
    # 测试代码
    dwg_file = os.path.join('outputs', '81206851-03_dwg.json')
    excel_file = os.path.join('outputs', '81206851-03_excel.json')
    appearance_map_file = os.path.join('maps', 'appearance_map.json')
    template_file = os.path.join('maps', 'report_map.xlsx')

    generate_template_report(dwg_file, excel_file, appearance_map_file, template_file)