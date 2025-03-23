#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import logging
from openpyxl import load_workbook
import xlrd
from typing import Dict, Any, Optional, List, Tuple

# 配置日志
logger = logging.getLogger(__name__)

def extract_cell_value(excel_file: str, cell_coord: str, sheet_index: int = 0) -> str:
    """
    从Excel文件中提取特定单元格的值，支持.xlsx和.xls格式
    
    Args:
        excel_file: Excel文件路径
        cell_coord: 单元格坐标，如'A1', 'C5'等
        sheet_index: 工作表索引，默认为第一个工作表
        
    Returns:
        str: 单元格的值，如果单元格为空或发生错误则返回空字符串
    """
    # 判断文件格式
    file_extension = os.path.splitext(excel_file)[1].lower()
    
    # 提取单元格坐标的行和列
    try:
        # 分割单元格坐标为列和行
        col_str = ''.join(filter(str.isalpha, cell_coord))
        row_str = ''.join(filter(str.isdigit, cell_coord))
        
        if not col_str or not row_str:
            logger.error(f"无效的单元格坐标: {cell_coord}")
            return ""
        
        # 将列字母转换为索引 (A->0, B->1, etc.)
        col_idx = 0
        for char in col_str:
            col_idx = col_idx * 26 + (ord(char.upper()) - ord('A')) + 1
        col_idx -= 1  # 转换为0-based索引
        
        # 将行字符串转换为整数，并减1变为0-based索引
        row_idx = int(row_str) - 1
    except Exception as e:
        logger.error(f"解析单元格坐标 {cell_coord} 时出错: {e}")
        return ""
    
    try:
        if file_extension == '.xlsx':
            # 使用openpyxl处理.xlsx文件
            wb = load_workbook(excel_file, data_only=True)
            
            # 获取工作表
            sheets = wb.worksheets
            if sheet_index >= len(sheets):
                logger.error(f"工作表索引 {sheet_index} 超出范围，工作簿中只有 {len(sheets)} 个工作表")
                return ""
            
            ws = sheets[sheet_index]
            
            # 获取单元格值
            cell_value = ws[cell_coord].value
            
            # 如果单元格值不为空，转换为字符串并返回
            if cell_value is not None:
                return str(cell_value).strip()
        
        elif file_extension == '.xls':
            # 使用xlrd处理.xls文件
            wb = xlrd.open_workbook(excel_file)
            
            # 获取工作表
            if sheet_index >= wb.nsheets:
                logger.error(f"工作表索引 {sheet_index} 超出范围，工作簿中只有 {wb.nsheets} 个工作表")
                return ""
            
            ws = wb.sheet_by_index(sheet_index)
            
            # 检查单元格坐标是否有效
            if row_idx >= ws.nrows or col_idx >= ws.ncols:
                logger.error(f"单元格坐标 {cell_coord} 超出范围")
                return ""
            
            # 获取单元格值
            cell_value = ws.cell_value(row_idx, col_idx)
            
            # 如果单元格值不为空，转换为字符串并返回
            if cell_value is not None and cell_value != '':
                return str(cell_value).strip()
        else:
            logger.error(f"不支持的文件格式: {file_extension}")
            return ""
        
        return ""
    except Exception as e:
        logger.error(f"从Excel文件 {excel_file} 中提取单元格 {cell_coord} 的值时出错: {e}")
        return ""

def extract_cells_from_excel(excel_file: str, cells: Dict[str, str], sheet_index: int = 0) -> Dict[str, str]:
    """
    从Excel文件中提取多个单元格的值，支持.xlsx和.xls格式
    
    Args:
        excel_file: Excel文件路径
        cells: 单元格映射，格式为 {'key': 'cell_coord', ...}，例如 {'product_code': 'C5', 'material': 'H6'}
        sheet_index: 工作表索引，默认为第一个工作表
        
    Returns:
        Dict[str, str]: 提取的单元格值，格式为 {'key': 'value', ...}
    """
    result = {}
    
    # 对每个单元格调用extract_cell_value函数
    for key, cell_coord in cells.items():
        cell_value = extract_cell_value(excel_file, cell_coord, sheet_index)
        result[key] = cell_value
    
    return result

def extract_product_info_direct(excel_file: str) -> Dict[str, str]:
    """
    直接从工艺流程卡Excel文件中提取产品信息
    
    Args:
        excel_file: Excel文件路径
        
    Returns:
        Dict[str, str]: 产品信息字典
    """
    # 定义要提取的单元格及其对应的键
    cells_to_extract = {
        "product_code": "D4",  # 产品编码在D4单元格
        "drawing_no": "C5",    # 图号在C5单元格
        "material": "H6",      # 材料在H6单元格
        "version": "A4",       # 版本号在A4单元格
        "part_name": "C6"      # 零件名称在C6单元格
    }
    
    # 提取单元格值
    result = extract_cells_from_excel(excel_file, cells_to_extract)
    
    # 处理产品编码，从"产品编码：000000"中提取"000000"
    if "product_code" in result and "：" in result["product_code"]:
        parts = result["product_code"].split("：", 1)
        if len(parts) > 1:
            result["product_code"] = parts[1].strip()
    elif "product_code" in result and ":" in result["product_code"]:
        parts = result["product_code"].split(":", 1)
        if len(parts) > 1:
            result["product_code"] = parts[1].strip()
    
    # 处理版本号，从"版本:03"中提取"03"
    if "version" in result and "：" in result["version"]:
        parts = result["version"].split("：", 1)
        if len(parts) > 1:
            result["version"] = parts[1].strip()
    elif "version" in result and ":" in result["version"]:
        parts = result["version"].split(":", 1)
        if len(parts) > 1:
            result["version"] = parts[1].strip()
    
    logger.info(f"直接从Excel文件提取的产品信息: {result}")
    return result

# 用于测试的代码
if __name__ == "__main__":
    # 假设工艺流程卡文件路径为test.xlsx
    excel_file = "test.xlsx"
    
    # 提取产品信息
    product_info = extract_product_info_direct(excel_file)
    print("产品信息:", product_info) 